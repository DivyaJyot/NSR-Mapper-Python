import datetime
import json
import re
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from config import config
from config import loggingImpl
from library import xlrd

#from library.xlwt import copy
#from dateutil.parser import parser
#from library.xlutils import filter
#from library.xlutils import copy

 
greenFill = PatternFill(start_color='00B050',
                                end_color='00B050', fill_type='solid')


class writer:
    
    """ 
    this class takes input the NSR template object name and tries to fill the information parsed from the parser
    

    planningWorkbookName = ""
    planningWorksheetName
    objectName = {}
    cols = {}
    validationRules = {}
    NokiaAttrMap = []
    worksheet = None
    workbook = None
    attributeRowNumber = 0
    validationRowNumber = 0
    outputWorkbookName = None
    """
   
    
    def __init__(self, planningSheetName, object_list=None ):
        self.planningSheetName = planningSheetName
       # self.jsonFileName = jsonFileName
        #self.planningWorksheetName = planningWorksheetName
        self.cols = {}
        self.validationRules = {}
        self.objectList = object_list
        self.write_logger=loggingImpl.loggingImpl.getLogger()
        self.write_logger.info("\n Start Writing Data to Excel--- ")
        
    def writeValues(self):
        self.openWorkBook()
    
    def openWorkBook(self):
         self.workbook = xlrd.open_workbook (self.planningSheetName)
         timestamp = time.strftime('%d-%m-%y_%H.%M.%S')
         self.write_logger.info(self.planningSheetName+ "--"+self.planningSheetName[self.planningSheetName.rfind("\\"):len(self.planningSheetName)])
         excelName=self.planningSheetName[self.planningSheetName.rfind("\\"):len(self.planningSheetName)-4]+timestamp
         self.outputWorkBookName=config.project_path+'/outputFiles/'+excelName+".xlsm"
         self.wb = load_workbook(filename=self.planningSheetName, keep_vba=True)
         self.wb.save(self.outputWorkBookName)
         self.wb= load_workbook(filename=self.outputWorkBookName,keep_vba=True)
         for objName in self.objectList:
             print("Writing Values for Object %s"%objName)
             #Open the corresponding worksheet and read header information
             self.openWorkSheet(objName)
             #FIll up the cols with respect to col number
             self.getCols()
             #Get values from the validation row
             self.getValidation()
             
             #Start writing value in the worksheet.
             self.writeObjectValues(objName)
             
         self.wb.save(self.outputWorkBookName)     
         os.startfile(self.outputWorkBookName)
    
        
    def openWorkSheet(self,worksheetName):    
         self.worksheet = self.workbook.sheet_by_name(worksheetName)
         self.attributeRowNumber=int(self.worksheet.cell_value(1,4))-1
         self.validationRowNumber=int(self.worksheet.cell_value(1,6))-1
         self.dataRowNumber=int(self.worksheet.cell_value(1,8))
         self.write_logger.info("attribute row-validation row---"+str(self.attributeRowNumber)+str(self.validationRowNumber))
         
         self.ws = self.wb[worksheetName]
         
    def getCols(self):
        for curr_col in range(1, self.worksheet.ncols):
             self.cols[self.worksheet.cell_value(self.attributeRowNumber, curr_col)] = curr_col
        #print (self.cols)
         
    def getValidation(self):
         for cur_col in range(1, self.worksheet.ncols):
             self.validationRules[self.worksheet.cell_value(self.attributeRowNumber, cur_col)] = str(self.worksheet.cell_value(self.validationRowNumber, cur_col)).split('\n')    
         #print (self.validationRules)
     
 
    
    def getValidationAsString(self,validationRule):
        return ",".join(str(item) for item in validationRule[1:] ) 
            
    def applyValidationToCell(self,ws,startingRow,nokiaAttr):
        #get Validation rule for this cell
        validationRule = self.validationRules[nokiaAttr]
        if "list:" in validationRule:
            validationString = self.getValidationAsString(validationRule)
            str = '"' + validationString + '"'
            dv = DataValidation(type="list",formula1=str,allow_blank=False, showDropDown=False)
            dv.error = 'Your entry is not as per the validation rule'
            dv.errorTitle = 'Invalid Entry'
            ws.add_data_validation(dv)
            cell = ws.cell(column=self.cols[nokiaAttr] + 1, row=startingRow)
 
            dv.add(cell)
            
                
    def writeObjectValues(self,objName):
        self.write_logger.info(self.wb.named_styles)
        if('style1' not in self.wb.named_styles):
            style1 = NamedStyle(name="style1")
            style1.font = Font(name='Calibri',size=8,color='FF000000') 
            style1.border = Border(left=Side(style='thin'),
							 right=Side(style='thin'),
							 top=Side(style='thin'),
							 bottom=Side(style='thin'))
						  
            style1.fill = PatternFill(start_color='BDD7EE',end_color='BDD7EE',fill_type='solid')
        else:
            style1=	self.wb._named_styles['style1']
						   
        if( 'style2' not in self.wb.named_styles):
            style2 = NamedStyle(name="style2")

            style2.font = Font(name='Calibri',
								   size=8,
								   color='FF000000') 
            style2.border = Border(left=Side(style='thin'),
							 right=Side(style='thin'),
							 top=Side(style='thin'),
							 bottom=Side(style='thin'))
            style2.fill = PatternFill(start_color='00B050',
						   end_color='00B050',
						   fill_type='solid')
        else:
            style2= self.wb._named_styles['style2']   
        username_json_path = config.project_path +'\\log\\'+config.customer+"\\"+objName+"\\"+objName+'_output.json' 
        with open(username_json_path,'r') as username_file:
            nokiaData = json.load(username_file)
        startingRow = self.dataRowNumber
         
        for nokiaRecord in nokiaData:
            cell = self.ws.cell(column=1,row=startingRow)
            try:
                cell.style = style2
            except:
                self.write_logger.debug("style already exist ")
            self.ws.cell(column=1,row=startingRow,value='ADD')
            for nokiaAttr in nokiaRecord:
                #self.write_logger.debug(self.validationRules[nokiaAttr])
                self.applyValidationToCell(self.ws,startingRow,nokiaAttr)
                #Create data validation object for this 
                cell = self.ws.cell(column=self.cols[nokiaAttr] + 1,row=startingRow)
                try:
                    cell.style = style1
                except:
                    self.write_logger.debug("style already exist ")
                
                self.ws.cell(column=self.cols[nokiaAttr] + 1,row=startingRow,value=nokiaRecord[nokiaAttr])
            startingRow = startingRow + 1
                #w_worksheet.write(startingRow, self.cols[nokiaAttr], nokiaRecord[nokiaAttr])
         
        
        self.write_logger.info("Finished writing Excel")         
     
